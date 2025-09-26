# plugins/export_excel.py (v1.1 - 修正 applymap 警告)

from plugin_interface import ExportPlugin
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re

# 定义一个正则表达式来匹配非法字符
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')


def sanitize_for_excel(value):
    """移除字符串中 Excel 不支持的非法字符"""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub(r'', value)
    return value


class ExcelExportPlugin(ExportPlugin):
    @property
    def name(self):
        return "导出为 Excel"

    @property
    def file_extension(self):
        return ".xlsx"

    @property
    def file_filter(self):
        return "Excel 工作簿 (*.xlsx)"

    @property
    def icon_name(self):
        return "excel"

    def export(self, data, file_path, header_text, log_callback):
        log_callback("  -> 开始生成 Excel 数据...")
        df = pd.DataFrame(data, columns=['类别', '品牌', '型号', '大小', '序列号', '生产日期', '保修查询链接'])

        # 使用新的 .map() 方法替代已弃用的 .applymap()
        df = df.map(sanitize_for_excel)

        start_row = 1 if header_text else 0

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='电脑配置信息', startrow=start_row)
            worksheet = writer.sheets['电脑配置信息']

            if header_text:
                log_callback("  -> 正在为 Excel 添加页眉文本...")
                worksheet.merge_cells(f'A1:{get_column_letter(df.shape[1])}1')
                header_cell = worksheet['A1']
                header_cell.value = header_text
                header_cell.font = Font(name='微软雅黑', size=14, bold=True, color="808080")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                log_callback("  -> Excel 页眉文本添加成功。")

            header_font = Font(bold=True)
            for cell in worksheet[start_row + 1]:
                cell.font = header_font

            alignment = Alignment(horizontal='left', vertical='center')
            for row_cells in worksheet.iter_rows(min_row=start_row + 1):
                for cell in row_cells:
                    cell.alignment = alignment
                    if isinstance(cell.value, str) and 'http' in cell.value:
                        cell.font = Font(color="0000FF", underline="single")
                        cell.hyperlink = cell.value

            for col_idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 60)  # 设置一个最大宽度上限

        log_callback("  -> Excel 格式化完成。")
        return file_path